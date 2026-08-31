#!/usr/bin/env python3
"""Update table values in a DIR stability report from vendor Data Statistics Excel.

Logic inference per table:
  1. Parse table CAPTION to determine: property, package, strength group range
  2. For each ROW: read storage condition (temp/humidity) from first cell
  3. Match against prior vendor data with same caption filters to INFER rounding rule
  4. Apply same aggregation (from caption) + storage condition + rounding rule to current vendor data

All changes use Word tracked changes (w:del/w:ins).

Usage:
    python update_tables.py <working_dir>
    python update_tables.py <working_dir> --packages "HDPE (125cc)"
    python update_tables.py <working_dir> --xml-dir /tmp/dir_xml --packages "HDPE (125cc)"
"""

import argparse
import glob
import math
import os
import re
import subprocess
import sys
import tempfile
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from shared_utils import find_max_id

TODAY = date.today().strftime('%Y-%m-%dT00:00:00Z')
AUTHOR = 'Claude'
UNPACK_SCRIPT = os.path.expanduser('~/.claude/skills/docx/scripts/office/unpack.py')
PACK_SCRIPT = os.path.expanduser('~/.claude/skills/docx/scripts/office/pack.py')


def _package_matches_filter(vendor_pkg_name, target_packages):
    """Check if a vendor data package name matches the user-specified target packages.

    Vendor data uses names like 'PCTFE', 'CFAF', 'Bottle', 'Bulk Simulator'.
    Users pass names like 'Blisters', 'Bottles', 'Bulk', 'CFAF'.
    This function does keyword matching to bridge the gap.
    """
    pkg_lower = vendor_pkg_name.lower()
    for target in target_packages:
        t_lower = target.lower()
        # Direct match (e.g., 'CFAF' in ['CFAF'])
        if t_lower == pkg_lower:
            return True
        # Keyword containment (e.g., 'pctfe' in 'pctfe blister', or 'bottle' in 'bottles')
        if t_lower in pkg_lower or pkg_lower in t_lower:
            return True
        # 'Blisters' should match 'PCTFE' and 'CFAF' (both are blister types)
        if 'blister' in t_lower and pkg_lower in ('pctfe', 'cfaf'):
            return True
        # 'Bottles' should match 'Bottle'
        if 'bottle' in t_lower and 'bottle' in pkg_lower:
            return True
        # 'Bulk' should match 'Bulk Simulator'
        if 'bulk' in t_lower and 'bulk' in pkg_lower:
            return True
    return False


def load_vendor_data(xlsx_path):
    """Load Data Statistics sheet from vendor Excel. Returns list of row dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if 'Data Statistics' not in wb.sheetnames:
        wb.close()
        return None

    ws = wb['Data Statistics']
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        data.append({headers[i]: row[i] for i in range(len(headers))})
    wb.close()
    return data


def _condition_matches(row, temp, humidity):
    """Check if a vendor data row matches the given temperature/humidity.

    Handles two formats:
      - Separate 'Temperature' and 'Humidity' columns (Analysis Data sheet)
      - Combined 'Stability_Condition' column like '30°C/75% RH' (Data Statistics sheet)
    """
    if 'Temperature' in row and 'Humidity' in row:
        return row.get('Temperature') == temp and row.get('Humidity') == humidity
    # Combined column: parse "30°C/75% RH" → (30, 75)
    cond = row.get('Stability_Condition', '')
    if not cond:
        return False
    m = re.match(r'(\d+)\s*°?\s*C\s*/\s*(\d+)\s*%', str(cond))
    if m:
        return int(m.group(1)) == temp and int(m.group(2)) == humidity
    return False


def get_vendor_values(vendor_data, prop, temp, humidity, stat_col, packages=None, strength_groups=None):
    """Get ALL matching values from vendor data (across strength groups)."""
    values = []
    for row in vendor_data:
        if row.get('Property') != prop:
            continue
        if not _condition_matches(row, temp, humidity):
            continue
        if packages and row.get('Package') not in packages:
            continue
        if strength_groups and row.get('Strength_Group') not in strength_groups:
            continue
        if stat_col in row and row[stat_col] is not None:
            try:
                values.append(float(row[stat_col]))
            except (ValueError, TypeError):
                continue
    return values


def extract_text(xml):
    """Extract all text from XML fragment."""
    return ''.join(re.findall(r'<w:t[^>]*>([^<]*)</w:t>', xml))


def extract_property(caption):
    """Identify the stability property from table caption text.

    Returns the property name as it appears in vendor data. Tries both
    'Dissolution X' and plain 'X' forms since naming varies between reports.
    """
    cap = caption.lower()
    if 'assay' in cap:
        return 'Assay (% Label Claim)'
    if 'water activity' in cap:
        return 'Aw'
    if 'degradation' in cap:
        return None
    dm = re.search(r'(average|minimum)\s+percent\s+dissolved.*?at\s+(\d+)\s+minutes', cap)
    if dm:
        prefix = 'Average' if dm.group(1) == 'average' else 'Minimum'
        # Return both possible forms — caller will try both against vendor data
        return f'{prefix} % Dissolved at {dm.group(2)} Minutes'
    return None


def find_property_in_vendor(prop, vendor_data):
    """Find the exact property name used in vendor data.

    Vendor data may use 'Dissolution Average % Dissolved...' or just 'Average % Dissolved...'
    This function tries both forms.
    """
    if not prop or not vendor_data:
        return prop

    # Check if prop exists directly
    for row in vendor_data:
        if row.get('Property') == prop:
            return prop

    # Try with 'Dissolution ' prefix
    with_prefix = f'Dissolution {prop}'
    for row in vendor_data:
        if row.get('Property') == with_prefix:
            return with_prefix

    # Try without 'Dissolution ' prefix
    if prop.startswith('Dissolution '):
        without_prefix = prop[len('Dissolution '):]
        for row in vendor_data:
            if row.get('Property') == without_prefix:
                return without_prefix

    return prop


def extract_package_from_caption(caption, vendor_data):
    """Match table caption to a package in vendor data.

    Strategy: extract all unique package names from vendor data, then check
    if any of them (or recognizable fragments) appear in the caption.
    Also checks if the caption contains words that are commonly used as
    synonyms for vendor package names.

    No hardcoded mapping — entirely data-driven.

    Returns (matched_package_name, matched_text) or (None, None).
    """
    cap_lower = caption.lower()

    # Get unique packages from vendor data
    vendor_packages = set()
    for row in vendor_data:
        pkg = row.get('Package')
        if pkg:
            vendor_packages.add(pkg)

    # Strategy 1: Direct match — check if any vendor package name appears in caption
    for pkg in vendor_packages:
        if pkg.lower() in cap_lower:
            return pkg, pkg.lower()

    # Strategy 2: Word-level matching — split package names into words
    # and check if significant words appear in caption
    # e.g., "30 Count Bottle" → check if "bottle" is in caption
    # e.g., "PCTFE Blister" → check if "pctfe" is in caption
    for pkg in vendor_packages:
        pkg_words = re.findall(r'[a-zA-Z]+', pkg.lower())
        # Use distinctive words (skip generic: count, group, etc.)
        skip_words = {'count', 'group', 'primary', 'stability', 'the', 'and', 'for', 'mg'}
        distinctive = [w for w in pkg_words if w not in skip_words and len(w) > 2]
        for word in distinctive:
            if word in cap_lower:
                return pkg, word

    # Strategy 3: If strategies 1 and 2 fail, return None.
    # The inference engine will brute-force try each vendor package individually
    # to find which one produces 100% cell matches — no synonym list needed.
    return None, None


def extract_strength_groups_from_caption(caption, vendor_data, prop):
    """Extract which strength groups the table covers based on caption text.

    Parses phrases like:
      "for the 1 – 12 mg Strengths" → groups containing doses 1-12 mg
      "for the 24 and 36 mg Dose Strengths" → groups containing 24 or 36 mg

    Returns list of matching Strength_Group values from vendor data, or None (all groups).
    """
    cap = caption.lower()

    # Look for dose range patterns
    # Pattern: "for the X – Y mg" or "for the X and Y mg" or "for the X - Y mg"
    range_match = re.search(r'for the\s+(\d+)\s*[–\-]\s*(\d+)\s*mg', cap)
    list_match = re.search(r'for the\s+([\d,\s]+(?:and\s+\d+)?)\s*mg', cap)

    if not range_match and not list_match:
        return None  # No strength filter in caption

    # Get target dose values
    target_doses = set()
    if range_match:
        low = int(range_match.group(1))
        high = int(range_match.group(2))
        target_doses = set(range(low, high + 1))
    elif list_match:
        dose_text = list_match.group(1)
        doses = re.findall(r'\d+', dose_text)
        target_doses = set(int(d) for d in doses)

    if not target_doses:
        return None

    # Find all unique strength groups in vendor data for this property
    all_groups = set()
    for row in vendor_data:
        if row.get('Property') == prop:
            sg = row.get('Strength_Group')
            if sg:
                all_groups.add(sg)

    if not all_groups:
        return None

    # Match groups that contain any of the target doses
    matching_groups = []
    for group in all_groups:
        # Extract dose numbers from group name like "Group 1: 1 mg & 3 mg"
        group_doses = set(int(d) for d in re.findall(r'(\d+)\s*mg', group))
        if group_doses & target_doses:
            matching_groups.append(group)

    return matching_groups if matching_groups else None


def parse_condition(text):
    """Parse storage condition (e.g., '30°C/65% RH') into (temp, humidity)."""
    m = re.match(r'(\d+).*?(\d+)%', text)
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def parse_value(text):
    """Parse a numeric table cell value. Returns (float_val, format_info) or (None, None)."""
    text = text.strip()
    if not text or text.startswith('NMT') or text.startswith('NLT') or text in ('N/A', '—', '-'):
        return None, None
    # Handle unicode minus (U+2212) which appears in some Word documents
    text = text.replace('−', '-')
    suffix = ''
    if text.endswith('%'):
        suffix = '%'
        text = text[:-1]
    try:
        val = float(text)
    except ValueError:
        return None, None
    decimals = len(text.split('.')[1]) if '.' in text else 0
    return val, {'decimals': decimals, 'suffix': suffix}


def apply_rounding(value, decimals, method):
    """Apply the specified rounding method."""
    factor = 10 ** decimals
    if method == 'floor':
        return math.floor(value * factor) / factor
    elif method == 'ceil':
        return math.ceil(value * factor) / factor
    else:  # 'round'
        return round(value, decimals)


def infer_rounding_for_cell(predicted_raw, cell_val, decimals):
    """Determine which rounding method converts predicted_raw to cell_val."""
    factor = 10 ** decimals
    candidates = {
        'round': round(predicted_raw, decimals),
        'floor': math.floor(predicted_raw * factor) / factor,
        'ceil': math.ceil(predicted_raw * factor) / factor,
    }
    for method, candidate in candidates.items():
        if abs(candidate - cell_val) < 0.01:
            return method
    return None


def infer_aggregation_and_rounding(table_cells, vendor_data, packages, strength_groups):
    """Infer both aggregation rule and rounding method for a table.

    Tests all combinations of aggregation (min/max/first/mean) × rounding (round/floor/ceil)
    for each stat column against the actual table values.

    Returns dict with rule info, or None if can't infer.
    """
    rule_funcs = {
        'min': lambda vals: min(vals),
        'max': lambda vals: max(vals),
        'first': lambda vals: vals[0],
        'mean': lambda vals: sum(vals) / len(vals),
    }

    best_result = None

    for min_agg in rule_funcs:
        for max_agg in rule_funcs:
            # For each aggregation combo, test all rounding combos
            for min_rnd in ('round', 'floor', 'ceil'):
                for max_rnd in ('round', 'floor', 'ceil'):
                    matches = 0
                    testable = 0

                    for prop, temp, humidity, stat_col, cell_val, fmt in table_cells:
                        values = get_vendor_values(
                            vendor_data, prop, temp, humidity, stat_col,
                            packages=packages, strength_groups=strength_groups)
                        if not values:
                            continue
                        testable += 1

                        agg = min_agg if stat_col == 'Minimum' else max_agg
                        rnd = min_rnd if stat_col == 'Minimum' else max_rnd

                        predicted_raw = rule_funcs[agg](values)
                        decimals = fmt.get('decimals', 1)
                        predicted = apply_rounding(predicted_raw, decimals, rnd)

                        if abs(predicted - cell_val) < 0.01:
                            matches += 1

                    if testable == 0:
                        continue

                    rate = matches / testable
                    if rate >= 1.0:
                        return {
                            'rule_min': min_agg,
                            'rule_max': max_agg,
                            'rounding_min': min_rnd,
                            'rounding_max': max_rnd,
                            'match_rate': rate,
                            'testable_cells': testable,
                            'confirmed': True,
                        }
                    if best_result is None or rate > best_result['match_rate']:
                        best_result = {
                            'rule_min': min_agg,
                            'rule_max': max_agg,
                            'rounding_min': min_rnd,
                            'rounding_max': max_rnd,
                            'match_rate': rate,
                            'testable_cells': testable,
                            'confirmed': False,
                        }

    return best_result


def update_tables(dir_work, current_data_xlsx, target_packages=None, skip_conditions=None,
                  cross_package=False):
    """Update table values in document.xml using per-table logic inference.

    For each table:
      1. Parse caption → property, package, strength groups
      2. For each row → storage condition (temp/humidity)
      3. Infer aggregation + rounding from prior vendor data
      4. Apply to current vendor data with same filters

    Returns count of applied changes.
    """
    current_vendor_data = load_vendor_data(current_data_xlsx)
    if current_vendor_data is None:
        print("  No 'Data Statistics' sheet found, skipping")
        return 0
    if not current_vendor_data:
        print("  Data Statistics sheet is empty")
        return 0

    # Load prior vendor data for inference
    working_dir = os.path.dirname(current_data_xlsx)
    prior_data_path = os.path.join(working_dir, 'prior_vendor', 'stability_plot_data.xlsx')
    prior_vendor_data = None
    if os.path.exists(prior_data_path):
        prior_vendor_data = load_vendor_data(prior_data_path)
        if prior_vendor_data:
            print(f"  Prior vendor data loaded ({len(prior_vendor_data)} rows)")
    else:
        print("  Prior vendor data not found — using adaptive defaults")

    doc_path = os.path.join(dir_work, 'word', 'document.xml')
    with open(doc_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find all data tables
    table_pattern = re.compile(r'<w:tbl>(.*?)</w:tbl>', re.DOTALL)
    table_rules = {}
    changes = []

    print("\n  Per-table logic inference:")

    for tbl_match in table_pattern.finditer(content):
        preceding = content[max(0, tbl_match.start() - 5000):tbl_match.start()]
        context = extract_text(preceding[-3000:])

        # Find table number
        table_num_match = re.search(r'Table\s+(\d+\.\d+[-–]\d+)', context)
        if not table_num_match:
            table_num_match = re.search(r'Table\s+(\d+\.\d+)', context)
        if not table_num_match:
            continue
        table_id = f"Table {table_num_match.group(1)}"

        # Get the full caption (text after "Table X.Y-Z.")
        caption_match = re.search(
            r'Table\s+' + re.escape(table_num_match.group(1)) + r'\.\s*(.*?)(?:Table \d|\Z)',
            context)
        caption = caption_match.group(1).strip()[:200] if caption_match else context[-200:]

        # 1. Extract property from caption
        prop = extract_property(caption)
        if not prop:
            continue

        # Resolve property name against vendor data (handles 'Dissolution' prefix variations)
        ref_data = prior_vendor_data or current_vendor_data
        prop = find_property_in_vendor(prop, ref_data)

        # 2. Extract package from caption (matched against vendor data's actual package names)
        prior_pkg, _ = extract_package_from_caption(caption, ref_data)

        # For applying to current data, find matching package in current vendor data
        current_pkg_name, _ = extract_package_from_caption(caption, current_vendor_data)

        # Package filter: skip tables that belong to non-target packages
        # In cross-package mode, bypass this — we intentionally fill non-matching sections
        if not cross_package:
            if target_packages and current_pkg_name:
                if not _package_matches_filter(current_pkg_name, target_packages):
                    continue

        # Also resolve property in current vendor data (may differ from prior)
        current_prop = find_property_in_vendor(prop, current_vendor_data)

        # 3. Extract strength groups from caption
        strength_groups = extract_strength_groups_from_caption(caption, ref_data, prop)

        # Parse table rows
        tbl_xml = tbl_match.group(1)
        row_pat = re.compile(r'<w:tr\b[^>]*>(.*?)</w:tr>', re.DOTALL)
        cell_pat = re.compile(r'<w:tc\b[^>]*>(.*?)</w:tc>', re.DOTALL)

        parsed_rows = []
        for rm in row_pat.finditer(tbl_xml):
            cells = [extract_text(cm.group(1)).strip() for cm in cell_pat.finditer(rm.group(1))]
            parsed_rows.append(cells)

        if len(parsed_rows) < 2 or len(parsed_rows[0]) < 2:
            continue

        hdrs = parsed_rows[0]

        # Collect cells: (prop, temp, humidity, stat_col, cell_val, fmt)
        table_cells = []
        for row_idx in range(1, len(parsed_rows)):
            row = parsed_rows[row_idx]
            if not row or not row[0]:
                continue
            temp, humidity = parse_condition(row[0])
            if temp is None:
                continue
            if skip_conditions and (temp, humidity) in skip_conditions:
                continue
            for col_idx in range(1, len(row)):
                val, fmt = parse_value(row[col_idx])
                if val is None:
                    continue
                col_hdr = hdrs[col_idx] if col_idx < len(hdrs) else ''
                if 'min' in col_hdr.lower():
                    stat_col = 'Minimum'
                elif 'max' in col_hdr.lower():
                    stat_col = 'Maximum'
                else:
                    continue
                table_cells.append((prop, temp, humidity, stat_col, val, fmt))

        if not table_cells:
            continue

        # 4. Infer aggregation + rounding from prior data
        rule = None
        if prior_vendor_data:
            # Try with caption-derived filters
            prior_packages = [prior_pkg] if prior_pkg else None
            rule = infer_aggregation_and_rounding(
                table_cells, prior_vendor_data, prior_packages, strength_groups)

            # If not confirmed, try without strength group filter
            if rule and not rule['confirmed'] and strength_groups:
                rule2 = infer_aggregation_and_rounding(
                    table_cells, prior_vendor_data, prior_packages, None)
                if rule2 and rule2['match_rate'] > rule['match_rate']:
                    rule = rule2

            # If still not confirmed, try all packages (no filter)
            if rule and not rule['confirmed']:
                rule3 = infer_aggregation_and_rounding(
                    table_cells, prior_vendor_data, None, strength_groups)
                if rule3 and rule3['match_rate'] > rule['match_rate']:
                    rule = rule3

            # If STILL not confirmed, try each individual package
            if rule and not rule['confirmed']:
                all_pkgs = set(row.get('Package') for row in prior_vendor_data if row.get('Package'))
                for pkg in sorted(all_pkgs):
                    rule4 = infer_aggregation_and_rounding(
                        table_cells, prior_vendor_data, [pkg], strength_groups)
                    if rule4 and rule4['confirmed']:
                        rule = rule4
                        prior_pkg = pkg  # Update for logging
                        break
                    if rule4 and rule4['match_rate'] > rule['match_rate']:
                        rule = rule4

        # Store rule
        if rule and rule['confirmed']:
            table_rules[table_id] = rule
            pkg_note = f" pkg={prior_pkg}" if prior_pkg else ""
            sg_note = f" sg={len(strength_groups)}grp" if strength_groups else ""
            print(f"    {table_id} [{prop[:25]}]:{pkg_note}{sg_note} "
                  f"agg=Min:{rule['rule_min']}/Max:{rule['rule_max']} "
                  f"rnd=Min:{rule['rounding_min']}/Max:{rule['rounding_max']} "
                  f"({rule['testable_cells']} cells) ✓")
        else:
            # Adaptive default
            table_rules[table_id] = {
                'rule_min': 'min', 'rule_max': 'max',
                'rounding_min': 'round', 'rounding_max': 'round',
                'confirmed': False,
                'match_rate': rule['match_rate'] if rule else 0,
            }
            rate_str = f"{rule['match_rate']:.0%}" if rule else "N/A"
            print(f"    {table_id} [{prop[:25]}]: best={rate_str} — using adaptive")

        # 5. Apply rule to current vendor data
        rule_info = table_rules[table_id]
        # In cross-package mode, always use target_packages for data lookup
        # (caption says "Bottles" but we want CFAF vendor data)
        if cross_package and target_packages:
            apply_packages = target_packages
        elif current_pkg_name:
            apply_packages = [current_pkg_name]
        else:
            apply_packages = target_packages

        # Resolve strength groups against current vendor data too
        current_strength_groups = extract_strength_groups_from_caption(caption, current_vendor_data, current_prop) if current_prop else strength_groups

        for prop_c, temp_c, humidity_c, stat_col_c, val_c, fmt_c in table_cells:
            agg_rule = rule_info['rule_min'] if stat_col_c == 'Minimum' else rule_info['rule_max']
            rnd_method = rule_info['rounding_min'] if stat_col_c == 'Minimum' else rule_info['rounding_max']

            # Get values from current vendor data with filters
            values = get_vendor_values(
                current_vendor_data, current_prop or prop_c, temp_c, humidity_c, stat_col_c,
                packages=apply_packages or target_packages,
                strength_groups=current_strength_groups)

            if not values:
                continue

            # Apply aggregation
            if agg_rule == 'min':
                raw_val = min(values)
            elif agg_rule == 'max':
                raw_val = max(values)
            elif agg_rule == 'mean':
                raw_val = sum(values) / len(values)
            else:
                raw_val = values[0]

            # Apply rounding
            decimals = fmt_c['decimals']
            new_val = apply_rounding(raw_val, decimals, rnd_method)
            new_fmt = f"{new_val:.{decimals}f}{fmt_c['suffix']}"
            old_fmt = f"{val_c:.{decimals}f}{fmt_c['suffix']}"

            if new_fmt != old_fmt:
                changes.append({'old': old_fmt, 'new': new_fmt, 'table': table_id})

    # Print summary
    confirmed_count = sum(1 for r in table_rules.values() if r.get('confirmed'))
    print(f"\n  Rules summary: {len(table_rules)} tables analyzed")
    print(f"  Confirmed (100%): {confirmed_count}, Adaptive fallback: {len(table_rules) - confirmed_count}")

    # Apply changes with tracked changes
    change_id = find_max_id(content) + 100
    applied = 0
    for ch in changes:
        escaped = re.escape(ch['old'])
        pattern = re.compile(
            r'(<w:r(?:\s[^>]*)?>)\s*'
            r'((?:<w:rPr>(?:(?!</w:rPr>).)*</w:rPr>\s*)?)'
            r'(<w:t(?:\s[^>]*)?>)(' + escaped + r')(</w:t>)\s*(</w:r>)',
            re.DOTALL)

        for m in pattern.finditer(content):
            prec = content[max(0, m.start() - 500):m.start()]
            if prec.rfind('<w:tc') > prec.rfind('</w:tc>'):
                del_id = change_id
                ins_id = change_id + 1
                change_id += 2
                rpr = m.group(2).strip()
                if not rpr:
                    rpr = ''
                replacement = (
                    f'<w:del w:id="{del_id}" w:author="{AUTHOR}" w:date="{TODAY}">'
                    f'<w:r>{rpr}<w:delText xml:space="preserve">{ch["old"]}</w:delText></w:r></w:del>'
                    f'<w:ins w:id="{ins_id}" w:author="{AUTHOR}" w:date="{TODAY}">'
                    f'<w:r>{rpr}<w:t xml:space="preserve">{ch["new"]}</w:t></w:r></w:ins>')
                content = content[:m.start()] + replacement + content[m.end():]
                applied += 1
                break

    with open(doc_path, 'w', encoding='utf-8') as f:
        f.write(content)

    return applied


def main():
    parser = argparse.ArgumentParser(description='Update table values in DIR report from vendor data')
    parser.add_argument('working_dir', help='Directory containing vendor data files')
    parser.add_argument('--xml-dir', default=None,
                        help='Pre-unpacked XML directory for the DIR (skip unpack/repack)')
    parser.add_argument('--packages', default=None,
                        help='Comma-separated package filter (e.g., "HDPE (125cc)")')
    parser.add_argument('--skip-conditions', default=None,
                        help='Comma-separated temp/humidity conditions to skip (e.g., "40/75")')
    parser.add_argument('--cross-package', action='store_true', default=False,
                        help='Cross-package mode: bypass caption-based package filter, '
                             'use target packages directly for vendor data lookup')
    args = parser.parse_args()

    working_dir = os.path.abspath(args.working_dir)
    target_packages = None
    if args.packages:
        target_packages = [p.strip() for p in args.packages.split(',')]
        print(f"Package filter: {target_packages}")
    if args.cross_package:
        print(f"Cross-package mode: using target packages for all table data lookup")

    skip_conditions = None
    if args.skip_conditions:
        skip_conditions = set()
        for cond in args.skip_conditions.split(','):
            parts = cond.strip().split('/')
            if len(parts) != 2:
                print(f"ERROR: Invalid condition format '{cond.strip()}' "
                      f"(expected 'temp/humidity', e.g. '40/75')", file=sys.stderr)
                sys.exit(1)
            try:
                skip_conditions.add((int(parts[0]), int(parts[1])))
            except ValueError:
                print(f"ERROR: Non-numeric condition '{cond.strip()}' "
                      f"(expected integers, e.g. '40/75')", file=sys.stderr)
                sys.exit(1)
        print(f"Skip conditions: {skip_conditions}")

    # Check required files
    current_data = os.path.join(working_dir, 'stability_plot_data.xlsx')
    if not os.path.exists(current_data):
        print("ERROR: stability_plot_data.xlsx not found in working dir", file=sys.stderr)
        sys.exit(1)

    print("Updating tables in DIR report")

    if args.xml_dir:
        tbl_count = update_tables(args.xml_dir, current_data, target_packages, skip_conditions,
                                   cross_package=args.cross_package)
    else:
        # Standalone: find and unpack DRAFT
        pattern = os.path.join(working_dir, 'DIR_Form_*_DRAFT.docx')
        drafts = glob.glob(pattern)
        if not drafts:
            print("ERROR: No DIR_Form_*_DRAFT.docx found", file=sys.stderr)
            sys.exit(1)
        docx_path = drafts[0]

        with tempfile.TemporaryDirectory(prefix='tbl_') as tmp_base:
            dir_work = os.path.join(tmp_base, 'dir')
            subprocess.run([sys.executable, UNPACK_SCRIPT, docx_path, dir_work],
                          capture_output=True, text=True, check=True)

            tbl_count = update_tables(dir_work, current_data, target_packages, skip_conditions,
                                       cross_package=args.cross_package)

            result = subprocess.run(
                [sys.executable, PACK_SCRIPT, dir_work, docx_path, '--validate', 'false'],
                capture_output=True, text=True)
            if result.returncode != 0:
                print(f"ERROR repacking: {result.stderr[:200]}", file=sys.stderr)
                sys.exit(1)

        print(f"Updated: {os.path.basename(docx_path)}")

    print(f"\n  Total table cells updated: {tbl_count}")


if __name__ == '__main__':
    main()
